# Libreria de manejo de JSON
import json

# Para leer el archivo JSON, con caracteres en espanol
with open('Catalogo_de_Bienes_Inmuebles_en_administracion_susceptibles_de_venta.json',  encoding='utf-8-sig') as myfile:
    data=myfile.read() # Lo gurada como objeto jason

# Lo castea a una lista
obj = json.loads(data)

lista = [] # Lista donde se guardaran los datos especificos

#Recorremos la lista con los datos
for x in obj:
	# Creamos un Diccionario con los datos NumeroSIAB , TipoInmueble y EntidadFederativa
	dic = dict(NumeroSIAB = x['NumeroSIAB'],TipoInmueble = x['TipoInmueble'],EntidadFederativa = x['EntidadFederativa'])
	lista.append(dic) # Se agrega a la lista el diccionario

# Importamos la funcion Workbook que nos crea un libro de excel
from openpyxl import Workbook
wb = Workbook() # creamos el libro de excel

# Activamos o creamos una hoja de calculo
ws = wb.active

# Ponemos los titulos e las celdas correspondiente [ A1, A2, B2 y C2 ]
ws['A1'] = 'Catalogo de Bienes Inmuebles en administracion susceptibles de venta' 
ws['A2'] = 'NumeroSIAB'
ws['B2'] = 'TipoInmueble'
ws['C2'] = 'EntidadFederativa'

# Recorremos la lista con los datos especificos
for casilla in lista:
	# Para cada casilla de la lista va a crear una lista con los datos NumeroSIAB, TipoInmueble y EntidadFederativa
	# Para luego agregar esa lista como fila en la hoja de calculo
	ws.append( [casilla['NumeroSIAB'],casilla['TipoInmueble'],casilla['EntidadFederativa']] )

# Finalmente guarda el archivo si no existe o lo sobreescribe si es que existe
wb.save("PuntoExtra.xlsx")